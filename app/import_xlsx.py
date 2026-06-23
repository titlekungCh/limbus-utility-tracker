"""One-time importer: read 'Limbus Utility Sheet.xlsx' and seed app/data.json.

Run once (or again to re-seed from the spreadsheet):
    python app/import_xlsx.py
After seeding, the app reads/writes app/data.json and the xlsx is no longer needed.
"""
import json
import os
import re
from datetime import datetime, date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "Limbus Utility Sheet.xlsx")
OUT = os.path.join(HERE, "data.json")


def clean(v):
    """Normalise a cell value for JSON."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        # The sheet uses U+FFFD / smart punctuation in a few names; keep as-is (UTF-8).
        return v.strip()
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def num(v, default=0):
    try:
        return clean(v) if isinstance(v, (int, float)) else (float(v) if v not in (None, "") else default)
    except (TypeError, ValueError):
        return default


def _hex(color):
    """openpyxl Color -> #RRGGBB (drops the alpha byte), or None."""
    rgb = getattr(color, "rgb", None) if color is not None else None
    return "#" + rgb[-6:] if isinstance(rgb, str) and len(rgb) in (6, 8) else None


def ifss_faction_colors(ws):
    """Faction colours from the sheet's conditional formatting on the Actual block
    (the NOT(ISERROR(SEARCH(("X"),...))) rules over G2:J13). Returns a list of
    {match, fill, font} in rule order (order = match priority). Picks up any new
    faction/source a future season introduces."""
    out = []
    for rng in ws.conditional_formatting:
        if "G2:J13" not in str(rng.sqref):
            continue
        for rule in ws.conditional_formatting[rng]:
            f = rule.formula[0] if rule.formula else ""
            m = re.search(r'SEARCH\(\("([^"]+)"\)', f)
            if not m or not rule.dxf or not rule.dxf.fill:
                continue
            fill = _hex(rule.dxf.fill.bgColor)
            font = _hex(rule.dxf.font.color) if (rule.dxf.font and rule.dxf.font.color) else None
            if fill:
                out.append({"match": m.group(1), "fill": fill, "font": font or "#202124"})
        break
    return out


def extract_ifss(wb):
    """Every 'IF SS<N> Sheet' -> structured page data, keyed by season number.
    above = Actual (F sinner, G:J) + Keyword (M:P), rows 2-13 (editable).
    stats = the sheet's pre-computed cards (read-only): predicted fingers A15:B19,
    actual fingers F15:G21, totals H15:I21, status counts L15:M21, combo legend
    N15:P21. faction = colours from the sheet's conditional formatting."""
    seasons = {}
    for name in wb.sheetnames:
        m = re.match(r"^IF SS(\d+) Sheet$", name)
        if not m:
            continue
        ws = wb[name]

        def v(coord):
            return clean(ws[coord].value)

        above = [{
            "sinner": v(f"F{r}"),
            "actual": [v(f"G{r}"), v(f"H{r}"), v(f"I{r}"), v(f"J{r}")],
            "keyword": [v(f"M{r}"), v(f"N{r}"), v(f"O{r}"), v(f"P{r}")],
        } for r in range(2, 14)]

        def rows(r0, r1, cols):
            return [[v(f"{c}{r}") for c in cols] for r in range(r0, r1 + 1)]

        seasons[m.group(1)] = {
            "season": int(m.group(1)),
            "above": above,
            "stats": {
                "predFinger": rows(15, 19, ["A", "B"]),
                "actFinger": rows(15, 21, ["F", "G"]),
                "totals": rows(15, 21, ["H", "I"]),
                "statusCounts": rows(15, 21, ["L", "M"]),
                "legend": rows(15, 21, ["N", "O", "P"]),
            },
            "faction": ifss_faction_colors(ws),
        }
    return seasons


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    inv = wb["Inventory"]
    ds = wb["DataSheet"]
    lun = wb["Lunacy"]

    def D(coord):
        return ds[coord].value

    def I(coord):
        return inv[coord].value

    # ---- 12 sinners: acronym (H1:H12), name (I1:I12), shards (J1:J12) ----
    sinners = []
    for r in range(1, 13):
        sinners.append({
            "name": clean(ds.cell(r, 9).value),       # I
            "acronym": clean(ds.cell(r, 8).value),    # H
            "shards": num(ds.cell(r, 10).value),      # J
            "cell": f"J{r}",
        })

    # ---- Manager XP curve: DataSheet D2:G201 (level, nextXP, maxEnk, maxIncrease) ----
    manager_curve = []
    for r in range(2, 202):
        lvl = ds.cell(r, 4).value  # D
        if lvl in (None, ""):
            continue
        manager_curve.append({
            "level": int(num(lvl)),
            "nextXP": num(ds.cell(r, 5).value),       # E = XP to reach next level
            "maxEnk": num(ds.cell(r, 6).value),       # F
            "maxIncrease": num(ds.cell(r, 7).value),  # G
        })

    # ---- Shard "needed" reference table: DataSheet L1:N12 (type -> need amounts) ----
    shard_table = []
    for r in range(1, 13):
        shard_table.append({
            "type": clean(ds.cell(r, 12).value),   # L
            "shardNeeded": num(ds.cell(r, 13).value),  # M
            "threadNeeded": num(ds.cell(r, 14).value),  # N
        })

    # ---- ID-level XP curve: DataSheet A2:C101 (level -> cumulative Total XP) ----
    id_level_curve = []
    for r in range(2, 102):
        lvl = ds.cell(r, 1).value  # A
        if lvl in (None, ""):
            continue
        id_level_curve.append({
            "level": int(num(lvl)),
            "totalXP": num(ds.cell(r, 3).value),  # C
        })

    state = {
        "meta": {"seededFrom": os.path.basename(XLSX), "version": 1},

        "inventory": {
            "crate": num(D("J13")),
            "pass": num(D("J14")),
            "threads": num(D("J15")),
            "tickets": {
                "I": num(D("J30")), "II": num(D("J31")),
                "III": num(D("J32")), "IV": num(D("J33")),
            },
        },

        "manager": {
            "level": int(num(I("J9"))),
            "currentXP": num(I("I8")),
            "nextLevelXP": num(I("I9")),
        },

        "lunacy": {
            "total": num(lun["A2"].value),
            "paid": num(lun["B2"].value),
            "extTickets": num(lun["D2"].value),
            "decaTickets": num(lun["E2"].value),
            "currentDate": clean(lun["G1"].value),
        },

        "md": {
            "dailyLeft": num(I("H18")),
            "weeklyLeft": num(I("I18")),
            "normalLeftTotal": num(I("J18")),
            "hard": [bool(I("H20")), bool(I("I20")), bool(I("J20"))],
            "normal": [bool(I("H22")), bool(I("I22")), bool(I("J22"))],
            # status theme per MD slot (H21:J21 hard, H23:J23 normal)
            "hardStatus": [clean(I("H21")), clean(I("I21")), clean(I("J21"))],
            "normalStatus": [clean(I("H23")), clean(I("I23")), clean(I("J23"))],
            "rental": bool(I("K20")),
            "rentalWeek": int(num(I("K22"))),
            # N17:N22 schedule checkboxes
            "schedule": [bool(inv.cell(r, 14).value) for r in range(17, 23)],
        },

        "currentDay": clean(I("K14")),
        "weekTilSeasonEnd": num(I("K18")),

        "uptie": {"sinner": clean(I("L18"))},
        "gacha": {"sinner": clean(I("K16")), "tier": clean(I("L16"))},

        # Intervallo event-shop planner (Inventory A1:G13 + reward type A15).
        #   consumables: cost/unit (row1), total qty (row2), bought (row9)
        #   rewards:     cost (row4), claimed (row11); ID/EGO cost depends on type
        "event": {
            "currency": num(I("F12")),               # F12 Current Currency (editable)
            "rewardType": clean(I("A15")),            # A15 Reward Type
            "items": [
                {
                    "name": clean(inv.cell(8, c).value),
                    "cost": num(inv.cell(1, c).value),
                    "total": num(inv.cell(2, c).value),
                    "bought": num(inv.cell(9, c).value),
                }
                for c in range(1, 8)  # A..G
            ],
            "rewards": [
                {
                    "name": clean(inv.cell(10, c).value),
                    "cost": num(inv.cell(4, c).value),
                    "claimed": bool(inv.cell(11, c).value),
                }
                for c in range(1, 7)  # A..F
            ],
        },

        # daily threads / xp constants pulled from DataSheet (used by logic)
        "constants": {
            "xpLux": num(D("J35")),            # 60
            "dailyManagerXP": num(D("J39")),   # 240
            "threadLuxSkip": num(D("J38")),    # 60
            "skipThread": num(D("J27")),       # 18  (Skip Thread 5; daily threads = *3)
            "dailyThreads": num(D("J28")),     # 54  (= skipThread*3; computed, read-only)
            # tickets gained per daily luxcavation, DataSheet H30:H33 (I,II,III,IV)
            "dailyLuxTickets": {
                "I": num(D("H30")), "II": num(D("H31")),
                "III": num(D("H32")), "IV": num(D("H33")),
            },
            "managerCurve": manager_curve,
            "shardTable": shard_table,
            "idLevelCurve": id_level_curve,
            # XP granted per EXP ticket tier (DataSheet K30:K33) — used by the
            # ID-leveling calculator's ticket breakdown.
            "ticketXP": {
                "I": num(D("K30")), "II": num(D("K31")),
                "III": num(D("K32")), "IV": num(D("K33")),
            },
        },

        "sinners": sinners,
    }

    # ---- Per-sinner shard PLAN (Inventory rows 24-32) ----
    # Faithful to the sheet's array formulas:
    #   A27 Shard Needed  = VLOOKUP(type, ShardNumber, 2)
    #   A28 Shard Owned   = VLOOKUP(sinner, SinnerShard, 2)  (live shards)
    #   A29 Crate Needed  = ((needed - owned)/2) * row24-toggle, floored at 0
    #   A32 Thread Needed = VLOOKUP(type, ShardNumber, 3)
    # The only editable inputs are the Shard Type (row 26) and the per-sinner
    # "enabled" toggle (row 24, B24:M24). Everything else is derived.
    shard_plan = []
    for idx, col in enumerate(range(2, 14)):  # B..M = sinners 0..11
        plan_main = clean(inv.cell(30, col).value)
        plan_sub = clean(inv.cell(31, col).value)
        target = " / ".join([p for p in (plan_main, plan_sub) if p])
        shard_plan.append({
            "type": clean(inv.cell(26, col).value),        # row 26 Shard Type
            "enabled": bool(inv.cell(24, col).value),       # row 24 include-in-crate toggle
            "target": target,                                # row 30/31 ID/EGO to shard
        })
    state["shardPlan"] = shard_plan

    # ---- ID Level list ----
    idl = wb["ID Level"]
    ids = []
    for r in range(2, idl.max_row + 1):
        name = idl.cell(r, 1).value  # A ID Name (group)
        full = idl.cell(r, 9).value  # I full name
        if not full and not name:
            continue
        # ID Tier is encoded as 1-3 repeats of a custom-font glyph (U+00D8);
        # 1 glyph = 1 star (max 3). Convert to a star count + ★ display string.
        tier_raw = idl.cell(r, 3).value
        stars = str(tier_raw).count("Ø") if tier_raw else 0
        ids.append({
            # name = col A only (e.g. "LCB Sinner"); the "[name] sinner" combined
            # form (col I) is just a formula, kept out of the editable data.
            "name": clean(name),
            "sinner": clean(idl.cell(r, 2).value),
            "tier": "★" * stars if stars else "",
            "tierStars": stars,
            "season": clean(idl.cell(r, 4).value),
            "keyword": clean(idl.cell(r, 5).value),
            "extraKeyword": clean(idl.cell(r, 13).value),  # col M
            "internalId": num(idl.cell(r, 6).value, None),
            "release": clean(idl.cell(r, 7).value),
            "acquired": bool(idl.cell(r, 8).value),
            "level": num(idl.cell(r, 10).value, None),
            "levelExtra": num(idl.cell(r, 11).value, 0),  # XP into next level
            "uptie": num(idl.cell(r, 12).value, None),
        })

    # ---- EGO Tier list ----
    egl = wb["EGO Tier"]
    egos = []
    for r in range(2, egl.max_row + 1):
        full = egl.cell(r, 10).value  # J full name
        short = egl.cell(r, 1).value
        if not full and not short:
            continue
        egos.append({
            # name = col A only (e.g. "Crow's Eye View"); col J is the "[name] sinner" formula.
            "name": clean(short),
            "sinner": clean(egl.cell(r, 2).value),
            "sin": clean(egl.cell(r, 3).value),
            "tier": clean(egl.cell(r, 4).value),
            "season": clean(egl.cell(r, 5).value),
            "keyword": clean(egl.cell(r, 6).value),         # col F
            "extraKeyword": clean(egl.cell(r, 12).value),   # col L
            "internalId": num(egl.cell(r, 7).value, None),
            "release": clean(egl.cell(r, 8).value),
            "acquired": bool(egl.cell(r, 9).value),
            "threadspin": num(egl.cell(r, 11).value, None),
        })

    # ---- grids: preserve row layout (blank rows kept; only trailing trimmed)
    #      so positional formulas (IF SS7 stats) map to fixed row indices. ----
    def grid(ws):
        rows = []
        for r in range(1, ws.max_row + 1):
            rows.append([clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)])
        while rows and all(v in (None, "") for v in rows[-1]):
            rows.pop()
        return rows

    teams = grid(wb["Bokgak Teams"])
    ifss7 = grid(wb["IF SS7 Sheet"])

    state["ids"] = ids
    state["egos"] = egos
    state["teams"] = teams
    state["ifss7"] = ifss7                # full grid (MD Teams page reads rows 22+)
    state["ifss"] = extract_ifss(wb)      # per-season structured IF SS pages

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

    print(f"Wrote {OUT}")
    print(f"  sinners={len(sinners)} ids={len(ids)} egos={len(egos)} "
          f"teams_rows={len(teams)} ifss7_rows={len(ifss7)} curve={len(manager_curve)}")


if __name__ == "__main__":
    main()
